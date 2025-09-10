"""
aria_members_final_v2.py

Scrapes all pages of https://aria.org.in/members-directory/
Writes a single Excel file (aria_members.xlsx) with columns:
type, name, company, mobile_no, email, website

Requirements:
    pip install selenium webdriver-manager beautifulsoup4 pandas openpyxl

Usage:
    Close aria_members.xlsx if it's open in Excel, then run:
    python aria_members_final_v2.py
"""

import time, re, os
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, ElementClickInterceptedException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

START_URL = "https://aria.org.in/members-directory/"
OUT_XLSX = Path("aria_members.xlsx")
HEADLESS = False
PAGE_WAIT = 1.2
CLICK_WAIT = 0.5

EMAIL_RE = re.compile(r'[\w\.-]+@[\w\.-]+\.\w+')
PHONE_RE = re.compile(r'(\+\d{1,3}[\s\-\.]?)?(\d{10,12})')
HTTP_RE = re.compile(r'https?://[^\s"\']+')

def setup_driver(headless=True):
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1600,1000")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)")
    options.add_experimental_option("excludeSwitches", ["enable-automation","enable-logging"])
    service = Service(ChromeDriverManager().install(), log_path="chromedriver.log")
    driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
    except Exception:
        pass
    return driver

def discover_max_page(driver):
    # look for any anchors with data-page attribute and numeric text
    nodes = driver.find_elements(By.CSS_SELECTOR, "a.page-link[data-page]")
    pages = []
    for n in nodes:
        dp = n.get_attribute("data-page")
        if dp and dp.isdigit():
            pages.append(int(dp))
    if pages:
        return max(pages)
    # fallback: try to parse last li >> a text
    try:
        last = driver.find_element(By.CSS_SELECTOR, "ul.pagination li:last-child a")
        txt = last.get_attribute("data-page") or last.text
        if txt and txt.isdigit():
            return int(txt)
    except Exception:
        pass
    # last resort: assume 14 (from earlier)
    return 14

def collect_pages_by_click(driver, start_url):
    driver.get(start_url)
    time.sleep(PAGE_WAIT)
    max_page = discover_max_page(driver)
    print("Detected max page:", max_page)

    pages_html = []
    for p in range(1, max_page+1):
        # if p == 1 we already loaded it
        if p == 1:
            print("Collecting page 1")
            pages_html.append(driver.page_source)
            continue
        # attempt to find element with data-page == p and click it
        print(f"Navigating to page {p} ...")
        tried = False
        try:
            # first try inside pagination container if present
            el = driver.find_element(By.CSS_SELECTOR, f"a.page-link[data-page='{p}']")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.12)
            try:
                el.click()
            except (ElementClickInterceptedException, StaleElementReferenceException):
                driver.execute_script("arguments[0].click();", el)
            time.sleep(PAGE_WAIT + CLICK_WAIT)
            pages_html.append(driver.page_source)
            tried = True
        except Exception as e:
            # fallback: try clicking by text (rare) or using JS to trigger the click
            try:
                driver.execute_script(
                    "var e = document.querySelector(\"a.page-link[data-page='%d']\"); if(e) e.click();" % p
                )
                time.sleep(PAGE_WAIT + CLICK_WAIT)
                pages_html.append(driver.page_source)
                tried = True
            except Exception:
                tried = False
        if not tried:
            print(f"Warning: Could not navigate to page {p}. Continuing.")
    print(f"Collected {len(pages_html)} pages (requested 1..{max_page}).")
    return pages_html

def parse_cards_from_html(html):
    soup = BeautifulSoup(html, "lxml")
    cards = soup.select("div.card.member-card")
    members = []
    for card in cards:
        typ = (card.select_one(".membercategory").get_text(strip=True) if card.select_one(".membercategory") else "").strip()
        name = (card.select_one(".itemtitle").get_text(strip=True) if card.select_one(".itemtitle") else "").strip()

        # company
        company = ""
        for li in card.select("ul.member-listgroup li.member-listgroup-item"):
            icon = li.select_one("i")
            if icon and "bi-briefcase" in " ".join(icon.get("class") or []):
                h = li.select_one("h6.title")
                if h:
                    company = h.get_text(strip=True)
                break

        # phone
        phone = ""
        for li in card.select("ul.member-listgroup li.member-listgroup-item"):
            icon = li.select_one("i")
            if icon and "bi-phone" in " ".join(icon.get("class") or []):
                h = li.select_one("h6.title")
                if h:
                    phone = h.get_text(strip=True)
                break
        phone = re.sub(r'\s+','', phone)  # remove internal spaces but preserve leading +

        # email
        email = ""
        for li in card.select("ul.member-listgroup li.member-listgroup-item"):
            icon = li.select_one("i")
            if icon and "bi-envelope" in " ".join(icon.get("class") or []):
                a = li.select_one("a[href^='mailto:']")
                if a:
                    email = a.get_text(strip=True)
                else:
                    h = li.select_one("h6.title")
                    if h:
                        email = h.get_text(strip=True)
                break

        # website
        website = ""
        for li in card.select("ul.member-listgroup li.member-listgroup-item"):
            icon = li.select_one("i")
            if icon and "bi-globe2" in " ".join(icon.get("class") or []):
                a = li.select_one("a[href^='http']")
                if a:
                    website = a.get("href").strip()
                break

        members.append({
            "type": typ,
            "name": name,
            "company": company,
            "mobile_no": phone,
            "email": email,
            "website": website
        })
    return members

def write_xlsx_only(members, out_xlsx=OUT_XLSX):
    # create dataframe and try to overwrite existing XLSX
    df = pd.DataFrame(members, columns=["type","name","company","mobile_no","email","website"])
    # attempt to remove existing file if present (if open in Excel, this will raise)
    if out_xlsx.exists():
        try:
            out_xlsx.unlink()
        except PermissionError:
            print(f"ERROR: Cannot overwrite {out_xlsx}. It appears to be open in another program (Excel).")
            print("Please close the file and re-run the script.")
            return False

    # write excel
    df.to_excel(out_xlsx, index=False, engine="openpyxl")
    # enforce Text format for mobile_no column
    wb = load_workbook(out_xlsx)
    ws = wb.active
    mobile_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if (cell.value or "").strip().lower() == "mobile_no":
            mobile_col_idx = idx
            break
    if mobile_col_idx:
        for r in range(1, ws.max_row+1):
            c = ws.cell(row=r, column=mobile_col_idx)
            c.number_format = '@'
            if c.value is None:
                continue
            if not isinstance(c.value, str):
                c.value = str(c.value)
    wb.save(out_xlsx)
    print(f"Wrote XLSX -> {out_xlsx} ({len(df)} rows).")
    return True

def main():
    driver = setup_driver(headless=HEADLESS)
    try:
        pages = collect_pages_by_click(driver, START_URL)
        all_members = []
        for i, html in enumerate(pages, start=1):
            print(f"Parsing page {i} ...")
            members = parse_cards_from_html(html)
            print(f" -> {len(members)} cards found")
            all_members.extend(members)
        # dedupe
        seen = set()
        dedup = []
        for m in all_members:
            key = (m.get("email","").lower(), re.sub(r'\D','', m.get("mobile_no","")), (m.get("name","") or "").lower())
            if key in seen:
                continue
            seen.add(key)
            dedup.append(m)
        print(f"Unique members collected: {len(dedup)}")
        ok = write_xlsx_only(dedup)
        if not ok:
            # user must close the file and re-run
            pass
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
